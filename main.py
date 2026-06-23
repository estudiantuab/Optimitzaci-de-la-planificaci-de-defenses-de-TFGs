import pandas as pd
import math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from dataclasses import dataclass, field
from typing import Annotated
from timefold.solver.domain import (
    planning_entity, planning_solution, PlanningId, PlanningVariable,
    ProblemFactCollectionProperty, PlanningEntityCollectionProperty,
    ValueRangeProvider, PlanningScore
)
from timefold.solver.score import (
    ConstraintFactory, Joiners, HardSoftScore, Constraint, constraint_provider
)
from timefold.solver import SolverFactory
from timefold.solver.config import (SolverConfig, ScoreDirectorFactoryConfig, TerminationConfig, Duration)
from timefold.solver.score import ConstraintCollectors
from timefold.solver import SolutionManager
from timefold.solver.score import HardMediumSoftScore
import sys

import pandas as pd
import math

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from dataclasses import dataclass, field
from typing import Annotated
from timefold.solver.domain import (
    planning_entity, planning_solution, PlanningId, PlanningVariable,
    ProblemFactCollectionProperty, PlanningEntityCollectionProperty,
    ValueRangeProvider, PlanningScore
)
from timefold.solver.score import (
    ConstraintFactory, Joiners, HardSoftScore, Constraint, constraint_provider
)
from timefold.solver import SolverFactory
from timefold.solver.config import (SolverConfig, ScoreDirectorFactoryConfig, TerminationConfig, Duration)
from timefold.solver.score import ConstraintCollectors
from timefold.solver import SolutionManager
from timefold.solver.score import HardMediumSoftScore

from lectura import generar_profes, analitzar_tutors_alumnes, crear_alumnes_i_excel, extreure_i_crear_sessions
from estructura import Tribunals, Profes, Alumnes, Aules, Sessions, Horari
from hard import c0, c1, c2, c2_bis, c3, c4, c5, c6, c7, c8, c9
from soft import r1, r2, r3, r4, r5

profes_comodi = ["profe_8", "profe_18", "profe_19", "profe_24", "profe_26", "profe_27", "profe_28", "profe_29", "profe_30", "profe_35"]

w2 = 15
w3 = 3

@constraint_provider
def define_constraints(constraint_factory: ConstraintFactory):
    return [
        #hard
        c0(constraint_factory),
        c1(constraint_factory),
        c2(constraint_factory),
        c3(constraint_factory),
        c4(constraint_factory),
        c5(constraint_factory),
        c6(constraint_factory),
        c7(constraint_factory),
        c8(constraint_factory),
        c9(constraint_factory, profes_comodi),
        #soft
        c2_bis(constraint_factory),
        r1(constraint_factory),
        r2(constraint_factory, w2),
        r3(constraint_factory, w3),
        r4(constraint_factory),
        r5(constraint_factory)
    ]

alumnes_setembre=[    "alumne_15",
    "alumne_46",
    "alumne_56",
    "alumne_35",
    "alumne_60",
    "alumne_40",
    "alumne_2",
    "alumne_26",
    "alumne_27"]

configuracio = [
    {'dia': 3, 'col_inici': 6, 'col_fi': 18, 'max_aules': 5},
    {'dia': 6, 'col_inici': 19, 'col_fi': 31, 'max_aules': 5},
    {'dia': 7, 'col_inici': 32, 'col_fi': 42, 'max_aules': 5}
]

excel_formulari = '0TFG-disponibilitat-i-preferencies-per-a-tribunals-2025_2026.xlsx'

df = pd.read_excel(excel_formulari)
profes=generar_profes('0tribunals-nombre-total-curs.xlsx')


df_alumnes_tutors= analitzar_tutors_alumnes(excel_formulari)
   
alumnes = crear_alumnes_i_excel(df_alumnes_tutors=df_alumnes_tutors, df_excepcions = pd.read_excel("0excepcions_arees_alumnes.xlsx"),llista_profes=profes, ruta_sortida_excel="llistat_alumnes.xlsx")
 
sessions = extreure_i_crear_sessions(excel_formulari, configuracio, "sessions.xlsx", alumnes_setembre)



from disponibilitat import omplir_i_exportar_disponibilitat


llista_profes_actualitzada = omplir_i_exportar_disponibilitat(
    ruta_arxiu_origen=excel_formulari, 
    llista_profes=profes, 
    llista_sessions=sessions, 
    configuracio_dies=configuracio
)

from voluntats import omplir_voluntats_profes

# Executem la funció que processa les dades
profes_processats = omplir_voluntats_profes( excel_formulari, profes, alumnes)

from disponibilitat import omplir_disponibilitat_alumnes


sessions_especials = {
    "alumne_9": ["sessio_1", "sessio_2", "sessio_3", "sessio_4", "sessio_5", "sessio_6", "sessio_7", "sessio_8", "sessio_9", "sessio_10", "sessio_11", "sessio_12", "sessio_13", "sessio_14", "sessio_15", "sessio_16", "sessio_17", "sessio_18", "sessio_19", "sessio_20", "sessio_21", "sessio_22", "sessio_23", "sessio_24", "sessio_25", "sessio_26"],
    "alumne_17": ["sessio_1", "sessio_2", "sessio_3", "sessio_4", "sessio_5", "sessio_6", "sessio_7", "sessio_8", "sessio_9", "sessio_10", "sessio_11", "sessio_12", "sessio_13", "sessio_14", "sessio_15", "sessio_16", "sessio_17", "sessio_18", "sessio_19", "sessio_20", "sessio_21", "sessio_22", "sessio_23", "sessio_24", "sessio_25", "sessio_26"],
    "alumne_25": ["sessio_1", "sessio_2", "sessio_3", "sessio_4", "sessio_5", "sessio_6", "sessio_7", "sessio_8", "sessio_9", "sessio_10", "sessio_11", "sessio_12", "sessio_13", "sessio_14", "sessio_15", "sessio_16", "sessio_17", "sessio_18", "sessio_19", "sessio_20"]
}

alumnes = omplir_disponibilitat_alumnes(alumnes, sessions, sessions_especials, alumnes_setembre)


from aules import generar_aules

aules = generar_aules(sessions)

# Creem els tribunals 
tribunals = [
    Tribunals(id=f"T{i+1}", alumne=alumne) 
    for i, alumne in enumerate(alumnes)
]



#for s in profes:
 #   print(s)

#for a in alumnes:
    #print(a)

#for se in sessions:
    #print(se)


# ## SOLUCIÓ

if len(sys.argv) < 3:
    print("ERROR: Has d'especificar el temps d'execució i l'identificador de la prova.")
    print("D'aquesta manera: python3 main.py <minuts> <identificador>")
    sys.exit(1)

minuts_limit = int(sys.argv[1])
identificador = sys.argv[2].strip()


solver_factory = SolverFactory.create(
    SolverConfig(
        solution_class=Horari,
        entity_class_list=[Tribunals],
        score_director_factory_config=ScoreDirectorFactoryConfig(
            constraint_provider_function=define_constraints
        ),
        termination_config=TerminationConfig(
            spent_limit=Duration(minutes=minuts_limit)
        )
    )
)


problem = Horari(sessions, aules, profes, alumnes, tribunals)


solver = solver_factory.build_solver()
solution = solver.solve(problem)
solution_manager = SolutionManager.create(solver_factory)
score_analysis = solution_manager.analyze(solution)

from generacio_excels import generar_excel, generar_excel_alumne

generar_excel(solution, minuts_limit, identificador)
generar_excel_alumne(solution, minuts_limit, identificador)

from generacio_excels_v2 import  generar_excel_analitic_complet_v2

generar_excel_analitic_complet_v2(solution, solver_factory, minuts_limit, identificador,
                                   score_analysis=score_analysis)