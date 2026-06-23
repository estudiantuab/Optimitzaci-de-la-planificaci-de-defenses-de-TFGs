import pandas as pd
import math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from dataclasses import dataclass, field
from typing import Annotated
from timefold.solver.domain import (
    planning_entity, planning_solution, PlanningId, PlanningVariable,
    ProblemFactCollectionProperty, PlanningEntityCollectionProperty,
    ValueRangeProvider, PlanningScore
)
from timefold.solver.score import (
    ConstraintFactory, Joiners, HardSoftScore, Constraint, constraint_provider
)
from timefold.solver import SolverFactory
from timefold.solver.config import (SolverConfig, ScoreDirectorFactoryConfig, TerminationConfig, Duration)
from timefold.solver.score import ConstraintCollectors
from timefold.solver import SolutionManager
from timefold.solver.score import HardMediumSoftScore
import sys

import pandas as pd
import math

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from dataclasses import dataclass, field
from typing import Annotated
from timefold.solver.domain import (
    planning_entity, planning_solution, PlanningId, PlanningVariable,
    ProblemFactCollectionProperty, PlanningEntityCollectionProperty,
    ValueRangeProvider, PlanningScore
)
from timefold.solver.score import (
    ConstraintFactory, Joiners, HardSoftScore, Constraint, constraint_provider
)
from timefold.solver import SolverFactory
from timefold.solver.config import (SolverConfig, ScoreDirectorFactoryConfig, TerminationConfig, Duration)
from timefold.solver.score import ConstraintCollectors
from timefold.solver import SolutionManager
from timefold.solver.score import HardMediumSoftScore




def generar_excel(solucio, minuts, label):
    dies = sorted(list(set(t.sessio.dia for t in solucio.tribunals if t.sessio)))
    hores = sorted(list(set(t.sessio.hora for t in solucio.tribunals if t.sessio)))
    aules_uniques = sorted(list(set(t.aula.id for t in solucio.tribunals if t.aula)))

    columnes_excel = []
    for aula in aules_uniques:
        columnes_excel.append(aula)
        columnes_excel.append(aula)

    nom_fitxer = f"{minuts}m_{label}_1_horari_solucio_professors.xlsx"
    
    with pd.ExcelWriter(nom_fitxer) as writer:
        for dia in dies:
            df = pd.DataFrame(index=hores, columns=range(len(columnes_excel)))
            
            for t in solucio.tribunals:
                if t.sessio and t.sessio.dia == dia and t.aula:
                    idx_aula = aules_uniques.index(t.aula.id) * 2
                    p1 = t.profe1.nom if t.profe1 else ""
                    p2 = t.profe2.nom if t.profe2 else ""
                    
                    df.iloc[hores.index(t.sessio.hora), idx_aula] = p1
                    df.iloc[hores.index(t.sessio.hora), idx_aula + 1] = p2
            
            df.columns = columnes_excel
            df.fillna("", inplace=True)
            df.to_excel(writer, sheet_name=f"Dia {dia}")

def generar_excel_alumne(solucio, minuts, label):
    dies = sorted(list(set(t.sessio.dia for t in solucio.tribunals if t.sessio)))
    hores = sorted(list(set(t.sessio.hora for t in solucio.tribunals if t.sessio)))
    aules = sorted(list(set(t.aula.id for t in solucio.tribunals if t.aula)))

    nom_fitxer = f"{minuts}m_{label}_2_horari_solucio_alumnes.xlsx"

    with pd.ExcelWriter(nom_fitxer) as writer:
        for dia in dies:
            df = pd.DataFrame(index=hores, columns=aules)

            for t in solucio.tribunals:
                if t.sessio and t.sessio.dia == dia and t.aula:
                    hora = t.sessio.hora
                    aula = t.aula.id
                    nom_alumne = t.alumne.nom if t.alumne else ""
                    df.at[hora, aula] = nom_alumne
            
            df.fillna("", inplace=True)
            df.to_excel(writer, sheet_name=f"Dia {dia}")
