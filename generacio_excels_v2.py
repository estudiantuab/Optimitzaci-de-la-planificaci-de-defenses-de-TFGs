import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from timefold.solver import SolutionManager

_VOL_LABEL = {0: "gens/tutor", 5: "indiferent", 15: "bastant", 20: "molt"}

def _vol_str(val):
    if val == "N/A" or val is None:
        return "N/A"
    return f"{val} ({_VOL_LABEL.get(val, '?')})"

def _area_ok(profe, alumne):
    if not profe or not alumne:
        return "N/A"
    return "✓" if profe.area_profe == alumne.area_alumne else "✗"

def _classifica_soft(nom_regla):
    nom = (nom_regla or "").lower()
    if "voluntat"   in nom:                          return "r1"
    if "àrea"       in nom or "discordança" in nom:  return "r2"
    if "aula"       in nom:                          return "r3"
    if "desplaçament" in nom or "dies" in nom:       return "r4"
    if "mortes"     in nom or "hores" in nom:        return "r5"
    return "c2bis_o_altre"

def _get_constraint_name(constraint_ref):
    for attr in ("constraint_id", "constraint_name"):
        val = getattr(constraint_ref, attr, None)
        if callable(val):
            try:
                val = val()
            except Exception:
                val = None
        if val:
            return str(val)
    return str(constraint_ref)

def _get_score_attr(score_obj, attr):
    val = getattr(score_obj, attr, 0)
    if callable(val):
        try:
            val = val()
        except Exception:
            val = 0
    return val or 0

def _build_styles():
    def fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def font(bold=False, color="000000", size=11, italic=False):
        return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)

    border = Border(
        left=Side(style="thin", color="D5D8DC"), right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),  bottom=Side(style="thin", color="D5D8DC"),
    )

    return {
        "fill_header":   fill("2C3E50"),
        "fill_zebra":    fill("F8F9F9"),
        "fill_fail":     fill("FADBD8"),
        "fill_success":  fill("D4EFDF"),
        "fill_warn":     fill("FEF9E7"),
        "fill_area_ok":  fill("D4EFDF"),
        "fill_area_nok": fill("FADBD8"),
        "fill_r1_pos":   fill("D4EFDF"),
        "fill_r_neg":    fill("FADBD8"),

        "font_title":     font(bold=True,  color="2C3E50", size=16),
        "font_header":    font(bold=True,  color="FFFFFF", size=11),
        "font_bold":      font(bold=True,  color="000000", size=11),
        "font_regular":   font(bold=False, color="000000", size=11),
        "font_ok":        font(bold=True,  color="145A32", size=11),
        "font_fail":      font(bold=True,  color="78281F", size=11),
        "font_italic_sm": font(bold=False, color="555555", size=10, italic=True),
        "border":         border,
    }



def generar_excel_analitic_complet_v2(solucio, solver_factory, minuts, label,
                                       score_analysis=None):
    solution_manager = SolutionManager.create(solver_factory)
    explanacio       = solution_manager.explain(solucio)
    indictment_map   = explanacio.indictment_map

    if score_analysis is None:
        score_analysis = solution_manager.analyze(solucio)

    nom_fitxer = f"{minuts}m_{label}_estudi_rendiment_solver.xlsx"
    st = _build_styles()

    files = []
    total_hard_err = 0
    total_soft_neg = 0

    for t in solucio.tribunals:
        hard_score = 0
        soft = {"r1": 0, "r2": 0, "r3": 0, "r4": 0, "r5": 0, "c2bis_o_altre": 0}
        regles_hard  = []
        regles_soft  = []

        if t in indictment_map:
            matches = (getattr(indictment_map[t], "constraint_matches", None)
                       or getattr(indictment_map[t], "constraint_match_set", []))
            for match in matches:
                nom_regla = getattr(match, "constraint_name", None)
                if not nom_regla and hasattr(match, "constraint_ref"):
                    nom_regla = _get_constraint_name(match.constraint_ref)
                nom_regla = nom_regla or "Desconeguda"

                p = match.score
                h = _get_score_attr(p, "hard_score")
                s = _get_score_attr(p, "soft_score")

                hard_score += h
                if h != 0:
                    regles_hard.append(f"{nom_regla}: {h}")
                if s != 0:
                    soft[_classifica_soft(nom_regla)] += s
                    regles_soft.append(f"{nom_regla}: {s}")

        soft_total_local = sum(soft.values())

        if hard_score < 0: total_hard_err += 1
        if soft_total_local < 0: total_soft_neg += 1

        sessio_id = t.sessio.id if t.sessio else None
        alumne_id = t.alumne.id if t.alumne else None
        tutor_id  = t.alumne.tutor_id if t.alumne else None

        def _p_disp(profe):
            if not profe or not sessio_id: return "N/A"
            return "SÍ" if profe.disponibilitat.get(sessio_id, 0) == 1 else "NO"

        def _p_vol_raw(profe):
            if not profe or not alumne_id: return "N/A"
            return profe.voluntats.get(alumne_id, 0)

        p1v = _p_vol_raw(t.profe1)
        p2v = _p_vol_raw(t.profe2)

        files.append({
            "ID_Tribunal":      t.id,
            "Alumne":           t.alumne.nom      if t.alumne  else "N/A",
            "Area_Alumne":      t.alumne.area_alumne if t.alumne else "N/A",
            "Disp_Alumne":      (t.alumne.disponibilitat_alumne.get(sessio_id, "N/A")
                                 if t.alumne and sessio_id else "N/A"),
            "Dia":              t.sessio.dia      if t.sessio  else -1,
            "Hora":             t.sessio.hora_id  if t.sessio  else "N/A",
            "Aula":             t.aula.id         if t.aula    else "N/A",

            "Profe_1":          t.profe1.nom      if t.profe1  else "N/A",
            "P1_Area":          t.profe1.area_profe if t.profe1 else "N/A",
            "P1_Area_ok":       _area_ok(t.profe1, t.alumne),
            "P1_Disp":          _p_disp(t.profe1),
            "P1_Voluntat":      _vol_str(p1v),
            "P1_Es_Tutor":      "SÍ" if t.profe1 and t.profe1.id == tutor_id else "NO",

            "Profe_2":          t.profe2.nom      if t.profe2  else "N/A",
            "P2_Area":          t.profe2.area_profe if t.profe2 else "N/A",
            "P2_Area_ok":       _area_ok(t.profe2, t.alumne),
            "P2_Disp":          _p_disp(t.profe2),
            "P2_Voluntat":      _vol_str(p2v),
            "P2_Es_Tutor":      "SÍ" if t.profe2 and t.profe2.id == tutor_id else "NO",
        })

    df_tribunals = pd.DataFrame(files)

    files_restr = []
    try:
        cm = score_analysis.constraint_map()
        items = cm.items() if hasattr(cm, "items") else []
        for cref, ca in items:
            nom  = _get_constraint_name(cref)
            scr  = ca.score
            h    = _get_score_attr(scr, "hard_score")
            me   = _get_score_attr(scr, "medium_score")
            s    = _get_score_attr(scr, "soft_score")
            mc   = getattr(ca, "match_count", 0) or 0
            if callable(mc):
                mc = mc()

            if   h  != 0: tipus, total = "Hard",   h
            elif me != 0: tipus, total = "Medium",  me
            else:         tipus, total = "Soft",    s

            avg = round(total / mc, 1) if mc > 0 else 0

            nota = ""
            cat  = _classifica_soft(nom)
            if cat in ("r3", "r5"):
                nota = "⚠ Valors per parella: pot semblar doble al desglossament per tribunal"
            elif cat == "r4":
                nota = "⚠ Per professor×dia: NO apareix a la columna per tribunal"

            files_restr.append({
                "Restricció":          nom,
                "Tipus":               tipus,
                "Score_Total":         total,
                "Num_Matches":         mc,
                "Score_Mig_per_Match": avg,
                "Nota":                nota,
            })
    except Exception as exc:
        files_restr = [{
            "Restricció": f"Error llegint constraint_map: {exc}",
            "Tipus": "-", "Score_Total": 0,
            "Num_Matches": 0, "Score_Mig_per_Match": 0, "Nota": "",
        }]

    files_restr.sort(key=lambda x: (0 if x["Tipus"] == "Hard" else 1 if x["Tipus"] == "Medium" else 2,
                                    x["Score_Total"]))
    df_restr = pd.DataFrame(files_restr)

    with pd.ExcelWriter(nom_fitxer, engine="openpyxl") as writer:
        df_restr.to_excel(    writer, sheet_name="Restriccions_Global", index=False)
        df_tribunals.to_excel(writer, sheet_name="Dades_Tribunals",     index=False)

        wb = writer.book

        ws_res = wb.create_sheet("Resum_Executiu", index=0)
        ws_res["A1"] = "INFORME ANALÍTIC DE RENDIMENT DEL SOLVER"
        ws_res["A1"].font = st["font_title"]

        for ci, txt in enumerate(["Mètrica", "Valor"], 1):
            c = ws_res.cell(row=3, column=ci, value=txt)
            c.font = st["font_header"]; c.fill = st["fill_header"]
            c.alignment = Alignment(horizontal="left", vertical="center")

        es_feasible = "SÍ" if solucio.puntuacio.hard_score == 0 else "NO (té conflictes HARD)"
        soft_global = solucio.puntuacio.soft_score

        metriques = [
            ("Label de la prova",            label),
            ("Temps de resolució",           f"{minuts} minuts"),
            ("Score HARD total",             solucio.puntuacio.hard_score),
            ("Score MEDIUM total",           getattr(solucio.puntuacio, "medium_score", 0)),
            ("Score SOFT total",             soft_global),
            ("Solució factible",             es_feasible),
            ("Total tribunals",              len(solucio.tribunals)),
            ("Tribunals amb error HARD",     total_hard_err),
            ("Tribunals amb soft negatiu",   total_soft_neg),
            ("─" * 30,                       "─" * 30),
            ("⚠ Nota r4 (viatges)",
             "El score de r4 és per professor×dia i NO apareix a la columna per tribunal. "
             "Consulta Restriccions_Global per al valor real."),
            ("⚠ Nota r3/r5 (aula/hores mortes)",
             "Són restriccions de parella: el valor per tribunal pot semblar el doble del real. "
             "Restriccions_Global té el total correcte sense doble compte."),
            ("⚠ Nota Soft_Total_Local",
             "Suma les restriccions que afecten directament el tribunal. "
             "No és igual al soft total global (que inclou r4 i efectes entre tribunals)."),
        ]

        for ri, (met, val) in enumerate(metriques, 4):
            c1 = ws_res.cell(row=ri, column=1, value=met)
            c2 = ws_res.cell(row=ri, column=2, value=val)
            c1.border = st["border"]; c2.border = st["border"]

            if "─" in str(met):
                c1.font = st["font_italic_sm"]; c2.font = st["font_italic_sm"]
                continue

            if "⚠" in str(met):
                c1.fill = st["fill_warn"];  c2.fill = st["fill_warn"]
                c1.font = st["font_italic_sm"]; c2.font = st["font_italic_sm"]
                continue

            c1.font = st["font_regular"]; c2.font = st["font_bold"]
            if ri % 2 == 1:
                c1.fill = st["fill_zebra"]; c2.fill = st["fill_zebra"]

            if met == "Solució factible":
                if not es_feasible.startswith("SÍ"):
                    c2.fill = st["fill_fail"];    c2.font = st["font_fail"]
                else:
                    c2.fill = st["fill_success"]; c2.font = st["font_ok"]
            elif met == "Score HARD total" and isinstance(val, (int, float)) and val < 0:
                c2.fill = st["fill_fail"]; c2.font = st["font_fail"]

        ws_res.column_dimensions["A"].width = 38
        ws_res.column_dimensions["B"].width = 80
        ws_res.merge_cells("A1:B1")

        ws_r = writer.sheets["Restriccions_Global"]
        ws_r.freeze_panes = "A2"

        for ci in range(1, len(df_restr.columns) + 1):
            c = ws_r.cell(row=1, column=ci)
            c.font = st["font_header"]; c.fill = st["fill_header"]
            c.alignment = Alignment(horizontal="center", vertical="center")

        for ri in range(2, len(df_restr) + 2):
            row_d = df_restr.iloc[ri - 2]
            total = row_d["Score_Total"]
            for ci in range(1, len(df_restr.columns) + 1):
                c   = ws_r.cell(row=ri, column=ci)
                col = df_restr.columns[ci - 1]
                c.border = st["border"]
                c.alignment = Alignment(
                    horizontal="right" if col in ("Score_Total", "Num_Matches", "Score_Mig_per_Match")
                    else "left"
                )
                if col == "Score_Total":
                    if total > 0:
                        c.fill = st["fill_r1_pos"]; c.font = st["font_ok"]
                    elif total < 0:
                        c.fill = st["fill_r_neg"];  c.font = st["font_fail"]
                    else:
                        c.font = st["font_regular"]
                elif ri % 2 == 0:
                    c.fill = st["fill_zebra"]

        for col in ws_r.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws_r.column_dimensions[col[0].column_letter].width = min(w + 4, 70)

        ws_d = writer.sheets["Dades_Tribunals"]
        ws_d.freeze_panes = "A2"
        ws_d.auto_filter.ref = f"A1:{get_column_letter(df_tribunals.shape[1])}1"

        cols_dreta   = {"Hard_Score", "Soft_r1_Voluntats", "Soft_r2_Àrees", "Soft_r3_Aula",
                        "Soft_r4_Viatges (*)", "Soft_r5_HoresMortes", "Soft_c2bis_o_Altre",
                        "Soft_Total_Local"}
        cols_centre  = {"Dia", "Hora", "Aula", "Disp_Alumne", "P1_Area_ok", "P2_Area_ok",
                        "P1_Disp", "P2_Disp", "P1_Es_Tutor", "P2_Es_Tutor"}

        for ci in range(1, df_tribunals.shape[1] + 1):
            c = ws_d.cell(row=1, column=ci)
            c.font = st["font_header"]; c.fill = st["fill_header"]
            c.alignment = Alignment(horizontal="center", vertical="center")

        for ri in range(2, len(df_tribunals) + 2):
            row_d    = df_tribunals.iloc[ri - 2]
            hard_val = row_d["Hard_Score"]
            is_even  = ri % 2 == 0

            for ci in range(1, df_tribunals.shape[1] + 1):
                c   = ws_d.cell(row=ri, column=ci)
                col = df_tribunals.columns[ci - 1]
                c.border = st["border"]
                c.font   = st["font_regular"]

                c.alignment = Alignment(
                    horizontal="right"  if col in cols_dreta  else
                    "center" if col in cols_centre else "left"
                )

                if hard_val < 0:
                    c.fill = st["fill_fail"]
                    if col in ("Hard_Score", "Regles_Hard"):
                        c.font = st["font_fail"]

                elif col == "P1_Area_ok":
                    val = str(row_d["P1_Area_ok"])
                    c.fill = st["fill_area_ok"] if val == "✓" else (st["fill_area_nok"] if val == "✗" else PatternFill())
                    c.font = st["font_ok"]      if val == "✓" else (st["font_fail"]      if val == "✗" else st["font_regular"])

                elif col == "P2_Area_ok":
                    val = str(row_d["P2_Area_ok"])
                    c.fill = st["fill_area_ok"] if val == "✓" else (st["fill_area_nok"] if val == "✗" else PatternFill())
                    c.font = st["font_ok"]      if val == "✓" else (st["font_fail"]      if val == "✗" else st["font_regular"])

                elif col == "Soft_r1_Voluntats" and row_d.get("Soft_r1_Voluntats", 0) > 0:
                    c.fill = st["fill_r1_pos"]

                elif col in ("Soft_r2_Àrees", "Soft_r3_Aula", "Soft_r5_HoresMortes"):
                    if row_d.get(col, 0) < 0:
                        c.fill = st["fill_r_neg"]
                    elif is_even:
                        c.fill = st["fill_zebra"]

                elif is_even:
                    c.fill = st["fill_zebra"]

        for col in ws_d.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws_d.column_dimensions[col[0].column_letter].width = min(w + 3, 45)